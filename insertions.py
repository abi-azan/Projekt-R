# SKRIPTA ZA GENEIRARANJE CSV-A S KOLONAMA PO ZELJI IZ data.xlsx

import openpyxl
import csv
from datetime import datetime

# Učitaj workbook
wb = openpyxl.load_workbook('data.xlsx')
ws = wb.active

rows = []

for row in range(2, 56):
    # Extrakciraj AnnounceDate iz kolone A
    ann_date_value = ws[f'A{row}'].value
    if isinstance(ann_date_value, str):
        ann_date_value = ann_date_value.rstrip('.')
        try:
            ann_date_parsed = datetime.strptime(ann_date_value, '%d.%m.%Y')
            ann_date_formatted = ann_date_parsed.strftime('%Y-%m-%d')
        except ValueError:
            ann_date_formatted = ann_date_value
    elif isinstance(ann_date_value, datetime):
        ann_date_formatted = ann_date_value.strftime('%Y-%m-%d')
    else:
        ann_date_formatted = ann_date_value

    # Extrakciraj EventDate iz kolone E
    date_value = ws[f'E{row}'].value
    if isinstance(date_value, str):
        # Ukloni točku na kraju ako postoji
        date_value = date_value.rstrip('.')
        # Parsiraj datum d.m.yyyy u datetime objekt
        try:
            date_parsed = datetime.strptime(date_value, '%d.%m.%Y')
            # Formatiraj u yyyy-mm-dd
            date_formatted = date_parsed.strftime('%Y-%m-%d')
        except ValueError:
            # Ako format nije kao očekivano, ostavi izvorni string
            date_formatted = date_value
    elif isinstance(date_value, datetime):
        date_formatted = date_value.strftime('%Y-%m-%d')
    else:
        date_formatted = date_value

    tickers_cell = ws[f'D{row}'].value
    if tickers_cell and isinstance(tickers_cell, str):
        tickers = [t.strip() for t in tickers_cell.split(';') if t.strip()]
        for ticker in tickers:
            rows.append({
                'Symbol': ticker, 
                'AnnDate': ann_date_formatted,
                'EventDate': date_formatted
            })

# Zapiši u CSV
with open('INSERTIONS_ANN_EVENT.csv', 'w', newline='', encoding='utf-8') as f:
    writer = csv.DictWriter(f, fieldnames=['Symbol', 'AnnDate', 'EventDate'])
    writer.writeheader()
    writer.writerows(rows)